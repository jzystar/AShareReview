#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A股数据获取和分析模块
主要用于每日收盘后的数据分析和短线交易指标计算
"""

import akshare as ak
import pandas as pd
import datetime
import warnings
import json
import os
import logging
from typing import Dict, List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

warnings.filterwarnings('ignore')

# 配置日志
def setup_logging():
    """配置日志系统"""
    # 创建logs目录
    os.makedirs('logs', exist_ok=True)
    
    # 配置日志格式
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s'
    
    # 配置根日志记录器
    logging.basicConfig(
        level=logging.INFO,
        format=log_format,
        handlers=[
            # 控制台输出
            logging.StreamHandler(),
            # 文件输出 - 按日期分割
            logging.FileHandler(
                f'logs/ashare_analysis_{datetime.datetime.now().strftime("%Y%m%d")}.log',
                encoding='utf-8'
            )
        ]
    )
    
    # 为akshare和pandas设置更高的日志级别，减少噪音
    logging.getLogger('akshare').setLevel(logging.WARNING)
    logging.getLogger('pandas').setLevel(logging.WARNING)
    logging.getLogger('urllib3').setLevel(logging.WARNING)

# 初始化日志
setup_logging()
logger = logging.getLogger(__name__)

class AShareAnalyzer:
    """A股数据分析器"""
    
    def __init__(self):
        logger.info("初始化AShareAnalyzer实例")
        self.today = datetime.datetime.now().strftime('%Y%m%d')
        self.yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y%m%d')
        self.data_dir = 'data'
        logger.info(f"设置分析日期: 今日={self.today}, 昨日={self.yesterday}")
        self.ensure_data_dir()
        
        # 板块价格涨跌幅限制
        self.price_limits = {
            '主板': 10.0,
            '科创板': 20.0,
            '创业板': 20.0,
            '北交所': 30.0
        }
        
        # 数据缓存
        self._cached_stock_data = None
        self._cache_timestamp = None
    
    def ensure_data_dir(self):
        """确保数据目录存在"""
        logger.debug(f"检查数据目录: {self.data_dir}")
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            logger.info(f"创建数据目录: {self.data_dir}")
            print(f"创建数据目录: {self.data_dir}")
        else:
            logger.debug(f"数据目录已存在: {self.data_dir}")
    
    def get_stock_data_with_board(self, force_refresh: bool = False) -> pd.DataFrame:
        """获取A股实时行情数据并添加板块信息（带缓存）"""
        logger.info(f"开始获取股票数据, force_refresh={force_refresh}")
        
        # 检查缓存是否有效（5分钟内）
        current_time = datetime.datetime.now()
        if (not force_refresh and 
            self._cached_stock_data is not None and 
            self._cache_timestamp is not None and 
            (current_time - self._cache_timestamp).seconds < 300):  # 5分钟缓存
            cache_age = (current_time - self._cache_timestamp).seconds
            logger.info(f"使用缓存数据，缓存年龄: {cache_age}秒")
            print("使用缓存的股票数据")
            return self._cached_stock_data
        
        try:
            logger.info("开始调用akshare接口获取实时股票数据")
            print("正在获取最新股票行情数据...")
            stock_data = ak.stock_zh_a_spot_em()
            
            if stock_data.empty:
                logger.warning("从akshare获取到空的股票数据")
                print("警告: 获取到空的股票数据")
                return pd.DataFrame()
            
            logger.info(f"成功从akshare获取 {len(stock_data)} 只股票的原始数据")
            
            # 添加板块分类
            logger.debug("开始为股票数据添加板块分类")
            stock_data['板块'] = stock_data['代码'].apply(self.classify_stock_board)
            
            # 统计各板块数量
            board_counts = stock_data['板块'].value_counts().to_dict()
            logger.info(f"板块分类统计: {board_counts}")
            
            # 更新缓存
            self._cached_stock_data = stock_data.copy()
            self._cache_timestamp = current_time
            logger.info(f"数据缓存已更新，时间戳: {self._cache_timestamp}")
            
            print(f"成功获取 {len(stock_data)} 只股票的行情数据")
            return stock_data
            
        except Exception as e:
            logger.error(f"获取股票数据失败: {e}", exc_info=True)
            print(f"获取股票数据失败: {e}")
            # 如果有缓存数据，返回缓存数据
            if self._cached_stock_data is not None:
                logger.warning("使用缓存数据作为备用")
                print("使用缓存数据作为备用")
                return self._cached_stock_data
            logger.error("无可用数据，返回空DataFrame")
            return pd.DataFrame()
    
    def save_results(self, results: Dict):
        """保存分析结果到JSON文件"""
        filename = f"{self.data_dir}/ashare_analysis_{self.today}.json"
        logger.info(f"开始保存分析结果到文件: {filename}")
        
        # 添加元数据
        data_to_save = {
            'date': self.today,
            'analysis_time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'results': results
        }
        
        # 统计结果数量
        result_count = len(results)
        total_items = sum(len(v) if isinstance(v, (list, dict)) else 1 for v in results.values())
        logger.info(f"准备保存 {result_count} 个分析结果，总计 {total_items} 个数据项")
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=2)
            
            # 检查文件大小
            file_size = os.path.getsize(filename)
            logger.info(f"数据保存成功: {filename}, 文件大小: {file_size/1024:.2f}KB")
            print(f"数据已保存到: {filename}")
            
        except Exception as e:
            logger.error(f"保存数据失败: {e}", exc_info=True)
            print(f"保存数据失败: {e}")
    
    def save_to_excel(self, results: Dict):
        """保存复盘数据到年度Excel文件"""
        year = datetime.datetime.now().year
        filename = f"复盘记录{year}.xlsx"
        logger.info(f"开始保存复盘数据到Excel文件: {filename}")
        
        try:
            # 提取需要的数据
            market_stats = results.get('市场统计', {})
            yesterday_perf = results.get('昨日涨停股表现', {})
            decline_analysis = results.get('第60个跌幅股票', 0)
            
            # 安全提取数值的函数
            def safe_extract_value(value, default=''):
                if value is None:
                    return default
                if isinstance(value, (int, float)):
                    return value
                if isinstance(value, str):
                    # 处理百分比格式
                    if '%' in value:
                        try:
                            return float(value.replace('%', ''))
                        except ValueError:
                            return value
                    # 处理"亿"格式
                    if '亿' in value:
                        try:
                            return float(value.replace('亿', ''))
                        except ValueError:
                            return value
                    return value
                return value
            
            # 准备行数据，按TODO.md中指定的顺序
            # 对于百分比数据，添加百分号（上证涨幅除外）
            def format_percentage(value, add_percent=True):
                if value == '' or value is None:
                    return ''
                if isinstance(value, (int, float)):
                    return f"{value}%" if add_percent else value
                return value
            
            # 格式化总成交额，添加"亿"字，数字部分使用整数
            def format_amount(value):
                if value == '' or value is None:
                    return ''
                # 如果已经包含"亿"字，提取数字部分转为整数再重新格式化
                if isinstance(value, str) and '亿' in value:
                    try:
                        num_part = float(value.replace('亿', ''))
                        return f"{int(num_part)}亿"
                    except ValueError:
                        return value
                # 如果是数字，转为整数后添加"亿"字
                if isinstance(value, (int, float)):
                    return f"{int(value)}亿"
                return value
            
            row_data = [
                datetime.datetime.now().strftime('%Y/%m/%d'),  # 1. 日期
                safe_extract_value(market_stats.get('上证量比', '')),  # 2. 上证量比
                safe_extract_value(market_stats.get('上证指数涨幅', '')),  # 3. 上证涨幅（不加百分号）
                format_amount(safe_extract_value(market_stats.get('总成交额', ''))),  # 4. 总成交额
                safe_extract_value(market_stats.get('涨跌停比', '')),  # 5. 涨跌停比
                format_percentage(safe_extract_value(market_stats.get('赚钱效应', ''))),  # 6. 赚钱效应
                format_percentage(safe_extract_value(market_stats.get('炸板率', ''))),  # 7. 炸板率
                safe_extract_value(decline_analysis),  # 8. 第60个跌幅股票
                safe_extract_value(market_stats.get('连板数量', '')),  # 9. 连板数量
                format_percentage(safe_extract_value(yesterday_perf.get('今日平均表现', ''))),  # 10. 昨日涨停表现
                format_percentage(safe_extract_value(yesterday_perf.get('昨日炸板股今日平均', '')))  # 11. 昨日炸板表现
            ]
            
            # 表头
            headers = [
                '日期', '上证量比', '上证涨幅', '总成交额', '涨跌停比', 
                '赚钱效应', '炸板率', '第60个跌幅股票', '连板数量', 
                '昨日涨停表现', '昨日炸板表现'
            ]
            
            # 提取赚钱效应数值，用于判断字体颜色
            money_effect_value = safe_extract_value(market_stats.get('赚钱效应', 0))
            if isinstance(money_effect_value, str) and '%' in money_effect_value:
                try:
                    money_effect_value = float(money_effect_value.replace('%', ''))
                except ValueError:
                    money_effect_value = 0
            elif not isinstance(money_effect_value, (int, float)):
                money_effect_value = 0
            
            # 检查文件是否存在
            new_row_num = None  # 记录新添加行的行号
            if os.path.exists(filename):
                logger.debug(f"Excel文件已存在，追加数据: {filename}")
                # 加载现有工作簿
                wb = load_workbook(filename)
                ws = wb.active
                
                # 检查是否已有今天的数据（避免重复）
                today_str = datetime.datetime.now().strftime('%Y/%m/%d')
                for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0] == today_str:
                        logger.warning(f"今天的数据已存在，跳过保存")
                        print(f"今天 ({today_str}) 的数据已存在于Excel文件中")
                        return
                        
                # 追加数据到最后一行
                ws.append(row_data)
                new_row_num = ws.max_row  # 获取新添加行的行号
            else:
                logger.debug(f"创建新的Excel文件: {filename}")
                # 创建新工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = f"{year}年复盘数据"
                
                # 添加表头
                ws.append(headers)
                # 添加数据
                ws.append(row_data)
                new_row_num = ws.max_row  # 数据在第2行（第1行是表头）
            
            # 根据赚钱效应设置字体颜色和对齐方式
            if new_row_num:
                font_color = "FF0000" if money_effect_value >= 50 else "00B050"  # 红色 或 绿色
                font = Font(color=font_color)
                alignment = Alignment(horizontal='left', vertical='top')  # 左对齐 + 上对齐
                logger.info(f"赚钱效应: {money_effect_value}%, 设置字体颜色: {'红色' if money_effect_value >= 50 else '绿色'}")
                
                # 设置所有列的对齐方式为左对齐
                for col in range(1, len(headers) + 1):  # 所有列
                    cell = ws.cell(row=new_row_num, column=col)
                    cell.alignment = alignment
                    
                    # 对第2列到第6列设置字体颜色
                    if 2 <= col <= 6:  # 列B(2)到F(6)
                        cell.font = font
            
            # 保存文件
            wb.save(filename)
            logger.info(f"Excel数据保存成功: {filename}")
            print(f"复盘数据已保存到Excel: {filename}")
            
        except Exception as e:
            logger.error(f"保存Excel数据失败: {e}", exc_info=True)
            print(f"保存Excel数据失败: {e}")
    
    def load_historical_data(self, date: str) -> Optional[Dict]:
        """加载指定日期的历史数据"""
        filename = f"{self.data_dir}/ashare_analysis_{date}.json"
        logger.info(f"尝试加载历史数据: {date}")
        
        if not os.path.exists(filename):
            # logger.warning(f"历史数据文件不存在: {filename}")
            # print(f"未找到日期 {date} 的数据文件")
            return None
        
        try:
            file_size = os.path.getsize(filename)
            logger.debug(f"开始读取文件: {filename}, 大小: {file_size/1024:.2f}KB")
            
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            logger.info(f"成功加载历史数据: {date}, 包含 {len(data.get('results', {}))} 个结果")
            return data
            
        except Exception as e:
            logger.error(f"加载历史数据失败: {e}", exc_info=True)
            print(f"加载历史数据失败: {e}")
            return None
    
    def get_historical_summary(self, max_days: int = 7, current_results: Optional[Dict] = None) -> Dict:
        """获取最近有数据的N天数据摘要（最多max_days天）"""
        summary = []
        base_date = datetime.datetime.now()
        today_str = base_date.strftime('%Y%m%d')
        
        # 向前搜索最多30天，找到有效数据
        for i in range(30):  # 最多搜索30天来找到有效数据
            if len(summary) >= max_days:  # 已经找到足够的数据
                break
                
            date = (base_date - datetime.timedelta(days=i)).strftime('%Y%m%d')
            
            # 对于今天的数据，优先使用传入的current_results，否则读取历史文件
            if date == today_str and current_results:
                logger.debug(f"使用当前运行结果处理今天的数据 {date}")
                data = {'results': current_results}
            else:
                data = self.load_historical_data(date)
            
            if data and 'results' in data:
                results = data['results']
                market_stats = results.get('市场统计', {})
                yesterday_perf = results.get('昨日涨停股表现', {})
                
                # 检查市场统计数据是否有效
                has_market_data = bool(market_stats and any(k in market_stats for k in ['涨停数量', '跌停数量', '赚钱效应']))
                
                # 检查昨日涨停表现数据是否有效
                has_yesterday_data = bool(yesterday_perf and '昨日涨停股数量' in yesterday_perf)
                
                # 只有当有有效数据时才添加到摘要中
                if has_market_data or has_yesterday_data:
                    # 安全提取数值的函数
                    def safe_extract_number(value, default=0):
                        if value is None:
                            return default
                        if isinstance(value, (int, float)):
                            return value
                        if isinstance(value, str):
                            # 处理百分比格式
                            if '%' in value:
                                try:
                                    return float(value.replace('%', ''))
                                except ValueError:
                                    return default
                            # 处理"亿"格式
                            if '亿' in value:
                                try:
                                    return float(value.replace('亿', ''))
                                except ValueError:
                                    return default
                            # 处理普通数字字符串
                            try:
                                return float(value)
                            except ValueError:
                                return default
                        return default
                    
                    day_summary = {
                        'date': date,
                        'limit_up_count': safe_extract_number(market_stats.get('涨停数量', 0)),
                        'limit_down_count': safe_extract_number(market_stats.get('跌停数量', 0)),
                        'up_down_ratio': market_stats.get('涨跌停比', 'N/A'),  # 涨跌停比
                        'total_amount': safe_extract_number(market_stats.get('总成交额', 0)),  # 总成交量
                        'sz_amount_rate': safe_extract_number(market_stats.get('上证量比', 0)),  # 上证量比
                        'sz_index_change': safe_extract_number(market_stats.get('上证指数涨幅', 0)),
                        'money_effect': safe_extract_number(market_stats.get('赚钱效应', 0)),
                        'exploded_rate': safe_extract_number(market_stats.get('炸板率', 0)),
                        'yesterday_limit_count': safe_extract_number(yesterday_perf.get('昨日涨停股数量', 0)),
                        'yesterday_avg_perf': safe_extract_number(yesterday_perf.get('今日平均表现', 0)),
                        'yesterday_up_ratio': safe_extract_number(yesterday_perf.get('今日上涨比例', 0)),
                        'exploded_avg_perf': safe_extract_number(yesterday_perf.get('昨日炸板股今日平均', 0)),  # 炸板表现
                        'has_valid_data': has_market_data and has_yesterday_data
                    }
                    summary.append(day_summary)
        
        logger.info(f"成功获取 {len(summary)} 天历史数据摘要")
        return {'historical_summary': summary}
        
    def classify_stock_board(self, stock_code: str) -> str:
        """根据股票代码判断所属板块"""
        code = str(stock_code)
        
        if code.startswith('68'):
            board = '科创板'
        elif code.startswith('3'):
            board = '创业板'
        elif code.startswith(('4', '8', '9')):
            board = '北交所'
        elif code.startswith(('0')):
            board = '主板'  # 深市主板
        elif code.startswith('6'):
            board = '主板'  # 沪市主板
        else:
            logger.debug(f"未知股票代码格式: {code}, 默认分类为主板")
            board = '主板'  # 默认归类为主板
        
        return board
    
    def get_stock_list_with_board(self) -> pd.DataFrame:
        """获取所有A股股票列表并标注板块"""
        try:
            # 获取所有A股股票
            stock_list = ak.stock_info_a_code_name()
            if not stock_list.empty:
                stock_list['板块'] = stock_list['code'].apply(self.classify_stock_board)
            return stock_list
        except Exception as e:
            print(f"获取股票列表失败: {e}")
            return pd.DataFrame()
    
    def get_stock_list(self) -> pd.DataFrame:
        """获取所有A股股票列表（包括主板、创业板、科创板、北交所）"""
        return self.get_stock_list_with_board()
    
    def get_daily_data(self, symbol: str) -> Optional[pd.DataFrame]:
        """获取单只股票的日K线数据"""
        try:
            df = ak.stock_zh_a_hist(symbol=symbol, period="daily", start_date="20240101", adjust="qfq")
            return df
        except Exception as e:
            print(f"获取股票 {symbol} 数据失败: {e}")
            return None
    
    def get_limit_up_stocks(self, min_days: int = 5) -> List[Dict]:
        """获取连续涨停天数大于等于min_days的股票"""
        logger.info(f"开始获取{min_days}连涨停板以上的股票")
        print(f"正在获取{min_days}连涨停板以上的股票...")
        
        try:
            logger.debug(f"调用akshare获取涨停股池数据, date={self.today}")
            # 获取涨停股票池
            limit_up_data = ak.stock_zt_pool_em(date=self.today)
            
            if limit_up_data.empty:
                logger.warning("没有获取到涨停股池数据")
                return []
            
            logger.info(f"获取到 {len(limit_up_data)} 只涨停股票")
            
            result = []
            consecutive_days_stats = {}

            
            for _, row in limit_up_data.iterrows():
                stock_code = row['代码']
                stock_name = row['名称']
                
                # 直接使用涨停股池中的连板数字段
                consecutive_days = row.get('连板数', 1)
                
                # 统计连板天数分布
                consecutive_days_stats[consecutive_days] = consecutive_days_stats.get(consecutive_days, 0) + 1
                
                if consecutive_days >= min_days:
                    result.append({
                        '代码': stock_code,
                        '名称': stock_name,
                        '连板天数': consecutive_days,
                        '最新价': row.get('最新价', 0),
                        '涨跌幅': row.get('涨跌幅', 0),
                        '封板资金': row.get('封板资金', 0),
                        '首次封板时间': row.get('首次封板时间', ''),
                        '炸板次数': row.get('炸板次数', 0)
                    })
            
            logger.info(f"连板天数分布: {consecutive_days_stats}")
            logger.info(f"筛选出 {len(result)} 只 {min_days}连板以上的股票")
            
            return result
            
        except Exception as e:
            logger.error(f"获取连续涨停股票失败: {e}", exc_info=True)
            print(f"获取连续涨停股票失败: {e}")
            return []
    
    
    def get_top_gainers_by_board(self, days: int, top_n: int = 5) -> Dict:
        """按板块获取指定天数内涨幅最大的前N只股票"""
        logger.info(f"开始按板块获取近{days}天涨幅前{top_n}名股票")
        print(f"正在按板块获取近{days}个交易日涨幅前{top_n}名股票...")
        
        try:
            # 使用缓存的股票数据
            stock_data = self.get_stock_data_with_board()
            
            if stock_data.empty:
                return {}
            
            results = {}
            
            # 按板块分别获取涨幅前N名
            for board in ['主板', '科创板', '创业板', '北交所']:
                logger.debug(f"开始处理{board}板块")
                board_data = stock_data[stock_data['板块'] == board]
                
                if board_data.empty:
                    logger.warning(f"{board}板块没有数据")
                    results[f'{board}涨幅前{top_n}'] = []
                    continue
                
                logger.debug(f"{board}板块共有 {len(board_data)} 只股票")
                # 按涨跌幅排序并取前N名
                board_data = board_data.sort_values('涨跌幅', ascending=False)
                
                top_gainers = []
                count = 0
                
                for _, row in board_data.iterrows():
                    if count >= top_n:
                        break
                        
                    top_gainers.append({
                        '代码': row['代码'],
                        '名称': row['名称'],
                        '板块': row['板块'],
                        '最新价': row['最新价'],
                        '涨跌幅': row['涨跌幅'],
                        '涨跌额': row['涨跌额'],
                        '成交量': row['成交量'],
                        '成交额': row['成交额'],
                        '振幅': row.get('振幅', 0),
                        '换手率': row.get('换手率', 0)
                    })
                    count += 1
                
                results[f'{board}涨幅前{top_n}'] = top_gainers
                logger.info(f"{board}板块成功获取 {len(top_gainers)} 只涨幅前{top_n}名股票")
            
            total_stocks = sum(len(stocks) for stocks in results.values())
            logger.info(f"所有板块处理完成，总计获取 {total_stocks} 只股票")
            return results
            
        except Exception as e:
            logger.error(f"获取涨幅排名失败: {e}", exc_info=True)
            print(f"获取涨幅排名失败: {e}")
            return {}
    
    def get_top_gainers(self, days: int, top_n: int = 5) -> List[Dict]:
        """获取指定天数内涨幅最大的前N只股票（保持向后兼容）"""
        board_results = self.get_top_gainers_by_board(days, top_n)
        # 合并所有板块的结果并按涨跌幅重新排序
        all_stocks = []
        for board_list in board_results.values():
            all_stocks.extend(board_list)
        
        # 按涨跌幅排序并取前N名
        all_stocks.sort(key=lambda x: x['涨跌幅'], reverse=True)
        return all_stocks[:top_n]
    
    def get_market_stats(self) -> Dict:
        """获取市场整体统计数据（按板块分类）"""
        logger.info("开始获取市场整体统计数据")
        print("正在获取市场整体数据...")
        
        try:
            logger.debug("获取上证指数数据")
            # 获取上证指数数据
            sz_index = ak.stock_zh_index_spot_em("沪深重要指数")
            
            # 使用缓存的股票数据
            stock_data = self.get_stock_data_with_board()
            
            if stock_data.empty:
                return {}
            
            # 计算各项指标
            total_amount = int(sz_index.iloc[0]['成交额'] + sz_index.iloc[1]['成交额']) // 10 ** 8
            sz_amount_rate = sz_index.iloc[0]['量比']
            
            # 上证指数涨幅
            sz_change = 0
            if not sz_index.empty:
                sz_change = sz_index.iloc[0]['涨跌幅']
                logger.info(f"上证指数当日涨跌幅: {sz_change}%")
            else:
                logger.warning("未能获取到上证指数数据")
            
            # 使用专门的涨停跌停接口获取准确数据
            limit_up_count = 0
            limit_down_count = 0
            exploded_count = 0
            
            # 按板块统计涨停跌停数据
            board_limit_stats = {}
            
            try:
                logger.debug(f"获取涨停股池数据, date={self.today}")
                # 获取涨停股池
                limit_up_data = ak.stock_zt_pool_em(date=self.today)
                if not limit_up_data.empty:
                    limit_up_data['板块'] = limit_up_data['代码'].apply(self.classify_stock_board)
                    limit_up_count = len(limit_up_data)
                    logger.info(f"获取到 {limit_up_count} 只涨停股票")
                    
                    # 按板块统计涨停数量
                    board_limit_up_counts = {}
                    greater_than_ten_of_limit_up = 0
                    for board in ['主板', '科创板', '创业板', '北交所']:
                        board_limit_up = limit_up_data[limit_up_data['板块'] == board]
                        count = len(board_limit_up)
                        board_limit_up_counts[board] = count
                        if board not in board_limit_stats:
                            board_limit_stats[board] = {}
                        board_limit_stats[board]['涨停数量'] = count
                        if board != '主板':
                            greater_than_ten_of_limit_up += count
                    
                    logger.info(f"各板块涨停数量: {board_limit_up_counts}")
                    logger.info(f"主板外涨停数量: {greater_than_ten_of_limit_up}")
                consecutive_limit_ups = len(limit_up_data[limit_up_data['连板数'] >= 2])
                logger.debug("获取炸板股池数据")
                # 获取炸板股池（开板的涨停股）
                exploded_data = ak.stock_zt_pool_zbgc_em(date=self.today)
                if not exploded_data.empty:
                    exploded_data['板块'] = exploded_data['代码'].apply(self.classify_stock_board)
                    exploded_count = len(exploded_data)
                    logger.info(f"获取到 {exploded_count} 只炸板股票")
                    
                    # 按板块统计炸板数量
                    board_exploded_counts = {}
                    for board in ['主板', '科创板', '创业板', '北交所']:
                        board_exploded = exploded_data[exploded_data['板块'] == board]
                        count = len(board_exploded)
                        board_exploded_counts[board] = count
                        if board not in board_limit_stats:
                            board_limit_stats[board] = {}
                        board_limit_stats[board]['炸板数量'] = count
                    
                    logger.info(f"各板块炸板数量: {board_exploded_counts}")
                
            except Exception as e:
                logger.error(f"获取涨停数据失败: {e}", exc_info=True)
                print(f"获取涨停数据失败: {e}")
            
            try:
                logger.debug("获取跌停股池数据")
                # 获取跌停股池
                limit_down_data = ak.stock_zt_pool_dtgc_em(date=self.today)
                if not limit_down_data.empty:
                    limit_down_data['板块'] = limit_down_data['代码'].apply(self.classify_stock_board)
                    limit_down_count = len(limit_down_data)
                    logger.info(f"获取到 {limit_down_count} 只跌停股票")
                    
                    # 按板块统计跌停数量
                    greater_than_ten_of_limit_down = 0
                    board_limit_down_counts = {}
                    for board in ['主板', '科创板', '创业板', '北交所']:
                        board_limit_down = limit_down_data[limit_down_data['板块'] == board]
                        count = len(board_limit_down)
                        board_limit_down_counts[board] = count
                        if board not in board_limit_stats:
                            board_limit_stats[board] = {}
                        board_limit_stats[board]['跌停数量'] = count
                        if board != '主板':
                            greater_than_ten_of_limit_down += count
                    
                    logger.info(f"各板块跌停数量: {board_limit_down_counts}")
                    logger.info(f"主板外跌停数量: {greater_than_ten_of_limit_down}")
            except Exception as e:
                logger.error(f"获取跌停数据失败: {e}", exc_info=True)
                print(f"获取跌停数据失败: {e}")
            
            # 按板块统计上涨下跌股票数量
            board_stats = {}
            greater_than_ten_of_not_limit_up = 0
            greater_than_ten_of_not_limit_down = 0
            for board in ['主板', '科创板', '创业板', '北交所']:
                board_data = stock_data[stock_data['板块'] == board]
                if not board_data.empty:
                    up_count = len(board_data[board_data['涨跌幅'] > 0])
                    down_count = len(board_data[board_data['涨跌幅'] < 0])
                    flat_count = len(board_data[board_data['涨跌幅'] == 0])
                    total_board_stocks = up_count + down_count + flat_count
                    money_effect = (up_count / total_board_stocks * 100) if total_board_stocks > 0 else 0
                    
                    board_stats[board] = {
                        '股票总数': total_board_stocks,
                        '上涨股票数': up_count,
                        '下跌股票数': down_count,
                        '平盘股票数': flat_count,
                        '赚钱效应': round(money_effect, 2)
                    }
                    if board in ['科创板', '创业板']:
                        greater_than_ten_of_not_limit_up += len(board_data[board_data['涨跌幅'] > 10]) - board_limit_stats[board]['涨停数量']
                        greater_than_ten_of_not_limit_down += len(board_data[board_data['涨跌幅'] < -10]) - board_limit_stats[board]['跌停数量']
                    # 添加涨停跌停数据
                    if board in board_limit_stats:
                        board_stats[board].update(board_limit_stats[board])
                        
                        # 计算炸板率
                        limit_up_board = board_limit_stats[board].get('涨停数量', 0)
                        exploded_board = board_limit_stats[board].get('炸板数量', 0)
                        exploded_rate = (exploded_board / (limit_up_board + exploded_board) * 100) if (limit_up_board + exploded_board) > 0 else 0
                        board_stats[board]['炸板率'] = round(exploded_rate, 2)
            
            # 整体数据
            up_count = len(stock_data[stock_data['涨跌幅'] > 0])
            down_count = len(stock_data[stock_data['涨跌幅'] < 0])
            flat_count = len(stock_data[stock_data['涨跌幅'] == 0])
            
            # 赚钱效应（上涨股票比例）
            total_stocks = up_count + down_count + flat_count
            money_effect = (up_count / total_stocks * 100) if total_stocks > 0 else 0
            
            logger.info(f"市场整体统计: 上涨{up_count}只, 下跌{down_count}只, 平盘{flat_count}只")
            logger.info(f"赚钱效应: {money_effect:.2f}%, 涨停{limit_up_count}只, 跌停{limit_down_count}只")
            
            # 计算整体炸板率
            exploded_rate = (exploded_count / (limit_up_count + exploded_count) * 100) if (limit_up_count + exploded_count) > 0 else 0
            if greater_than_ten_of_limit_down > 0:
                up_down_rate = f"{limit_up_count}({greater_than_ten_of_limit_up})+{greater_than_ten_of_not_limit_up}:{limit_down_count}({greater_than_ten_of_limit_down})+{greater_than_ten_of_not_limit_down}"
            else:
                up_down_rate = f"{limit_up_count}({greater_than_ten_of_limit_up})+{greater_than_ten_of_not_limit_up}:{limit_down_count}+{greater_than_ten_of_not_limit_down}"
            result = {
                '总成交额': f'{total_amount}亿',
                '上证指数涨幅': sz_change,
                '上证量比': sz_amount_rate,
                '涨跌停比': up_down_rate,
                '涨停数量': limit_up_count,
                '主板外涨停数量': greater_than_ten_of_limit_up,
                '涨幅大于10%的非涨停股票数': greater_than_ten_of_not_limit_up,
                '跌停数量': limit_down_count,
                '主板外跌停数量': greater_than_ten_of_limit_down,
                '跌幅大于10%的非跌停股票数': greater_than_ten_of_not_limit_down,
                '连板数量': consecutive_limit_ups,
                '上涨股票数': up_count,
                '下跌股票数': down_count,
                '平盘股票数': flat_count,
                '赚钱效应': round(money_effect, 2),
                '炸板率': round(exploded_rate, 2),
                
                '分板块统计': board_stats
                
            }
            
            return result
            
        except Exception as e:
            logger.error(f"获取市场统计数据失败: {e}", exc_info=True)
            print(f"获取市场统计数据失败: {e}")
            return {}
    
    def get_yesterday_performance(self) -> Dict:
        """获取昨日涨停股今日表现"""
        print("正在分析昨日涨停股表现...")
        
        try:
            # 使用专门的昨日涨停股池接口
            yesterday_limit_up = ak.stock_zt_pool_previous_em(date=self.today)
            
            if yesterday_limit_up.empty:
                return {'昨日涨停股表现': '无数据'}
            
            # 使用缓存的股票数据
            today_data = self.get_stock_data_with_board()
            today_dict = {row['代码']: row for _, row in today_data.iterrows()}
            
            performance = []
            for _, stock in yesterday_limit_up.iterrows():
                code = stock['代码']
                if code in today_dict:
                    today_stock = today_dict[code]
                    performance.append({
                        '代码': code,
                        '名称': stock['名称'],
                        '今日涨幅': today_stock['涨跌幅'],
                        '昨日连板数': stock.get('昨日连板数', 1)
                    })
            
            if performance:
                df = pd.DataFrame(performance)
                avg_performance = df['今日涨幅'].mean()
                up_ratio = len(df[df['今日涨幅'] > 0]) / len(df) * 100
                
                # 获取炸板股表现
                try:
                    exploded_data = ak.stock_zt_pool_zbgc_em(date=self.yesterday)
                    exploded_performance = []
                    for _, stock in exploded_data.iterrows():
                        code = stock['代码']
                        if code in today_dict:
                            today_stock = today_dict[code]
                            exploded_performance.append(today_stock['涨跌幅'])
                    
                    exploded_avg = sum(exploded_performance) / len(exploded_performance) if exploded_performance else 0
                except:
                    exploded_avg = 0
                
                return {
                    '昨日涨停股数量': len(performance),
                    '今日平均表现': round(avg_performance, 2),
                    '今日上涨比例': round(up_ratio, 2),
                    '昨日炸板股今日平均': round(exploded_avg, 2)
                }
            
            return {'昨日涨停股表现': '无有效数据'}
            
        except Exception as e:
            print(f"分析昨日涨停股表现失败: {e}")
            return {'昨日涨停股表现': f'分析失败: {e}'}
    
    def get_intraday_extremes(self) -> Dict:
        """获取当日盘中极值股票（按板块分类统计，设置不同阈值）"""
        print("正在获取当日盘中极值股票...")
        
        try:
            # 使用缓存的股票数据
            stock_data = self.get_stock_data_with_board()
            
            if stock_data.empty:
                return {}
            
            # 计算收盘相对最低点涨幅和相对最高点跌幅
            stock_data['收盘较最低涨幅'] = ((stock_data['最新价'] - stock_data['最低']) / stock_data['最低'] * 100).round(2)
            stock_data['收盘较最高跌幅'] = ((stock_data['最新价'] - stock_data['最高']) / stock_data['最高'] * 100).round(2)
            
            # 设置不同板块的阈值
            board_thresholds = {
                '主板': 15.0,
                '科创板': 25.0,
                '创业板': 25.0,
                '北交所': 35.0
            }
            
            results = {}
            
            # 按板块分别统计极值
            for board in ['主板', '科创板', '创业板', '北交所']:
                board_data = stock_data[stock_data['板块'] == board]
                
                if board_data.empty:
                    continue
                
                threshold = board_thresholds[board]
                
                # 收盘比当天最低点涨幅大于阈值的股票，取前5个
                filtered_gainers = board_data[board_data['收盘较最低涨幅'] > threshold]
                top_low_gainers = filtered_gainers.nlargest(5, '收盘较最低涨幅')[
                    ['代码', '名称', '最新价', '最低', '收盘较最低涨幅', '板块']
                ].to_dict('records')
                
                # 收盘比当天最高点跌幅大于阈值的股票（绝对值），取前5个
                # 对于跌幅，我们看绝对值大于阈值的（即跌幅超过阈值）
                filtered_losers = board_data[board_data['收盘较最高跌幅'] < -threshold]
                top_high_losers = filtered_losers.nsmallest(5, '收盘较最高跌幅')[
                    ['代码', '名称', '最新价', '最高', '收盘较最高跌幅', '板块']
                ].to_dict('records')
                
                results[f'{board}-收盘较最低涨幅前5(>{threshold}%)'] = top_low_gainers
                results[f'{board}-收盘较最高跌幅前5(>{threshold}%)'] = top_high_losers
            
            return results
            
        except Exception as e:
            print(f"获取盘中极值股票失败: {e}")
            return {}
    
    def get_sector_analysis(self) -> Dict:
        """获取行业板块分析"""
        print("正在分析行业板块表现...")
        
        try:
            # 获取行业板块数据
            sector_data = ak.stock_board_industry_name_em()
            
            if sector_data.empty:
                return {}
            
            # 按涨跌幅排序
            sector_data = sector_data.sort_values('涨跌幅', ascending=False)
            
            # 涨幅最大的前5个板块
            top_sectors = sector_data.head(5)[
                ['板块名称', '涨跌幅', '总市值', '换手率']
            ].to_dict('records')
            
            return {
                '涨幅最大板块': top_sectors
            }
            
        except Exception as e:
            print(f"获取板块分析失败: {e}")
            return {}
    
    def get_decline_analysis(self) -> Dict:
        """获取跌幅分析"""
        print("正在分析跌幅数据...")
        
        try:
            # 使用缓存的股票数据
            stock_data = self.get_stock_data_with_board()
            
            if stock_data.empty:
                return {}
            
            # 按跌幅排序
            declining_stocks = stock_data[stock_data['涨跌幅'] < 0].sort_values('涨跌幅')
            
            # 第60个跌幅最大的股票
            decline_60th = None
            if len(declining_stocks) >= 60:
                decline_60th =  declining_stocks.iloc[59]['涨跌幅']
                
            
            # 近3天跌幅超过20%的股票（简化版本，仅用当日跌幅估算）
            #heavy_decline_count = len(stock_data[stock_data['涨跌幅'] <= -15])  # 用单日跌幅估算
            
            return {
                '第60个跌幅股票': decline_60th,
            }
            
        except Exception as e:
            print(f"获取跌幅分析失败: {e}")
            return {}
    
    def run_analysis(self) -> Dict:
        """运行完整分析"""
        logger.info(f"="*50)
        logger.info(f"开始A股市场全面分析 - {self.today}")
        logger.info(f"="*50)
        print("="*50)
        print(f"开始A股市场分析 - {self.today}")
        print("="*50)
        
        results = {}
        analysis_start_time = datetime.datetime.now()
        
        # 1. 5连涨停板以上股票
        logger.info("步骤1: 获取5连涨停板以上股票")
        limit_up_stocks = self.get_limit_up_stocks(5)
        results['5连涨停以上股票'] = limit_up_stocks
        logger.info(f"步骤1完成: 找到 {len(limit_up_stocks)} 只5连板以上股票")
        
        # 2. 市场整体数据
        logger.info("步骤2: 获取市场整体统计数据")
        market_stats = self.get_market_stats()
        results['市场统计'] = market_stats
        logger.info("步骤2完成: 市场统计数据获取完毕")
        
        # 3. 昨日涨停股表现
        logger.info("步骤4: 分析昨日涨停股表现")
        yesterday_perf = self.get_yesterday_performance()
        results['昨日涨停股表现'] = yesterday_perf
        logger.info("步骤3完成: 昨日涨停股表现分析完毕")
        
        # 4. 盘中极值股票
        logger.info("步骤5: 获取盘中极值股票")
        intraday_extremes = self.get_intraday_extremes()
        results.update(intraday_extremes)
        logger.info(f"步骤4完成: 获取了 {len(intraday_extremes)} 项极值数据")
        
        # 5. 板块分析
        logger.info("步骤6: 进行行业板块分析")
        sector_analysis = self.get_sector_analysis()
        results.update(sector_analysis)
        logger.info(f"步骤5完成: 获取了 {len(sector_analysis)} 项板块分析数据")
        
        # 6. 跌幅分析
        logger.info("步骤7: 进行跌幅分析")
        decline_analysis = self.get_decline_analysis()
        results.update(decline_analysis)
        logger.info(f"步骤6完成: 获取了 {len(decline_analysis)} 项跌幅分析数据")
        
        # 7. 获取历史数据摘要（包含今天的数据）
        logger.info("步骤8: 获取历史数据摘要")
        historical_summary = self.get_historical_summary(7, results)
        results.update(historical_summary)
        logger.info(f"步骤7完成: 获取了 {len(historical_summary)} 项历史数据")
        
        # 统计分析用时
        analysis_end_time = datetime.datetime.now()
        analysis_duration = (analysis_end_time - analysis_start_time).total_seconds()
        
        total_results = len(results)
        logger.info(f"全部分析完成! 总计 {total_results} 项结果, 耗时 {analysis_duration:.2f} 秒")
        
        return results

def main():
    """主函数"""
    logger.info("程序启动 - A股市场分析系统")
    
    try:
        analyzer = AShareAnalyzer()
        logger.info("分析器初始化成功，开始执行分析")
        
        results = analyzer.run_analysis()
        
        # 保存结果到文件
        logger.info("开始保存分析结果")
        analyzer.save_results(results)
        logger.info("分析结果保存成功")
        
        # 保存到Excel文件
        logger.info("开始保存复盘数据到Excel")
        analyzer.save_to_excel(results)
        logger.info("复盘数据Excel保存成功")
        
        # 输出结果
        logger.info("开始输出分析结果汇总")
        print("\n" + "="*50)
        print("分析结果汇总")
        print("="*50)
        
        for key, value in results.items():
            # 跳过历史数据摘要的显示，太长了
            if key == 'historical_summary':
                continue
                
            print(f"\n【{key}】")
            logger.debug(f"显示结果: {key}")
            
            if isinstance(value, list):
                for i, item in enumerate(value, 1):
                    print(f"  {i}. {item}")
            elif isinstance(value, dict):
                for k, v in value.items():
                    print(f"  {k}: {v}")
            else:
                print(f"  {value}")
        
        # 显示历史数据摘要
        if 'historical_summary' in results:
            historical = results['historical_summary']
            if historical:
                logger.info(f"显示近7天历史数据摘要，包含 {len(historical)} 天数据")
                print(f"\n【近{len(historical)}日市场概况】")
                print("日期       涨停 跌停 涨跌停比          成交额 上证涨幅   量比  赚钱效应 炸板率 | 昨涨停数 涨停表现  上涨率 炸板表现")
                print("-" * 130)
                for day in historical:
                    # 检查数据有效性
                    valid_marker = "" if day.get('has_valid_data', False) else " *"
                    up_down_ratio = str(day.get('up_down_ratio', 'N/A'))
                    if len(up_down_ratio) > 18:
                        up_down_ratio = up_down_ratio[:16] + ".."
                    
                    print(f"{day['date']} {day['limit_up_count']:4d} {day['limit_down_count']:4d} "
                          f"{up_down_ratio:18s} {day['total_amount']:6.0f}亿 {day['sz_index_change']:7.2f}% {day['sz_amount_rate']:5.2f} "
                          f"{day['money_effect']:7.2f}% {day['exploded_rate']:6.2f}%  | "
                          f"{day['yesterday_limit_count']:7d} {day['yesterday_avg_perf']:8.2f}% {day['yesterday_up_ratio']:6.2f}% {day['exploded_avg_perf']:7.2f}%{valid_marker}")
                
                # 添加说明
                if any(not day.get('has_valid_data', False) for day in historical):
                    print("\n* 标记的日期数据不完整")
                    
                # 添加字段说明
                print("\n字段说明:")
                print("涨跌停比=涨停(主板外)+涨幅>10%非涨停:跌停(主板外)+跌幅>10%非跌停")
                print("炸板表现=昨日炸板股今日平均表现")
            else:
                logger.warning("历史数据摘要为空")
        
        logger.info("程序执行成功完成")
        
    except KeyboardInterrupt:
        logger.warning("程序被用户中断")
        print("\n程序被用户中断")
    except Exception as e:
        logger.error(f"程序执行过程中出现错误: {e}", exc_info=True)
        print(f"程序执行失败: {e}")
        raise
    finally:
        logger.info("程序结束")

if __name__ == "__main__":
    main()